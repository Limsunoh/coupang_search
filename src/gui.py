"""Tkinter 기반 GUI."""

from datetime import datetime
import threading
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from selenium.common.exceptions import (
    InvalidSessionIdException,
    TimeoutException,
    WebDriverException,
)

from src.scraper import CoupangScraper, KeywordResult


class App(tk.Tk):
    MAX_KEYWORDS = 100
    TTUDUNG_SORT_ORDER = {"넙덕": 0, "뼈다구": 1, "돼지": 2}

    def __init__(self) -> None:
        super().__init__()
        self.title("Coupang Keyword Analyzer")
        self.geometry("1250x650")
        self.minsize(1200, 600)

        # 쿠팡이 헤드리스를 차단하므로 창이 뜨는 모드로 사용한다.
        self.scraper = CoupangScraper(headless=False)
        self._is_running = False
        self.results: List[KeywordResult] = []
        self.errors: List[str] = []

        self._build_widgets()
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    def _build_widgets(self) -> None:
        top = ttk.Frame(self, padding=10)
        top.pack(fill="x")

        ttk.Label(
            top,
            text=f"키워드 (한 줄에 하나씩, 최대 {self.MAX_KEYWORDS}개)",
        ).pack(anchor="w")

        self.text_input = tk.Text(top, height=8)
        self.text_input.pack(fill="x", pady=(4, 8))

        button_frame = ttk.Frame(top)
        button_frame.pack(anchor="e")

        self.save_button = ttk.Button(
            button_frame,
            text="엑셀 저장",
            command=self.on_save,
            state="disabled",
        )
        self.save_button.pack(side="left", padx=(0, 6))

        self.run_button = ttk.Button(
            button_frame,
            text="분석 시작",
            command=self.on_run,
        )
        self.run_button.pack(side="left")

        self.status_var = tk.StringVar(value="대기 중")
        ttk.Label(self, textvariable=self.status_var, padding=(10, 0)).pack(anchor="w")

        # 결과 테이블 + 스크롤바
        table_frame = ttk.Frame(self)
        table_frame.pack(fill="both", expand=True, padx=10, pady=10)

        columns = (
            "keyword",
            "ttudung",
            "total",
            "rocket",
            "seller_rocket",
            "rocket_ratio",
            "seller_rocket_ratio",
            "seller_rocket_review_over_2000",
            "seller_rocket_max_review_count",
            "seller_rocket_review_pass",
            "rocket_ratio_pass",
        )
        headings = {
            "keyword": "키워드",
            "ttudung": "뚜둥",
            "total": "전체",
            "rocket": "로켓",
            "seller_rocket": "판매자로켓",
            "rocket_ratio": "로켓 %",
            "seller_rocket_ratio": "판매자로켓 %",
            "seller_rocket_review_over_2000": "댓글 2000+",
            "seller_rocket_max_review_count": "최대 댓글",
            "seller_rocket_review_pass": "댓글 기준",
            "rocket_ratio_pass": "로켓 기준",
        }
        widths = {
            "keyword": 200,
            "ttudung": 70,
            "total": 70,
            "rocket": 70,
            "seller_rocket": 70,
            "rocket_ratio": 90,
            "seller_rocket_ratio": 90,
            "seller_rocket_review_over_2000": 90,
            "seller_rocket_max_review_count": 90,
            "seller_rocket_review_pass": 70,
            "rocket_ratio_pass": 70,
        }

        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings")
        for col in columns:
            anchor = "w" if col == "keyword" else "center"
            self.tree.heading(col, text=headings[col])
            self.tree.column(col, width=widths[col], anchor=anchor)
        self.tree.tag_configure("error", background="#FFE5E5")

        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)

        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

    def on_run(self) -> None:
        if self._is_running:
            return

        raw = self.text_input.get("1.0", "end").strip()
        keywords = [line.strip() for line in raw.splitlines() if line.strip()]

        if not keywords:
            messagebox.showwarning("입력 필요", "키워드를 1개 이상 입력하세요.")
            return
        if len(keywords) > self.MAX_KEYWORDS:
            messagebox.showwarning(
                "제한 초과",
                f"키워드는 최대 {self.MAX_KEYWORDS}개까지 입력할 수 있습니다.",
            )
            return

        self._is_running = True
        self.run_button.config(state="disabled")
        self.save_button.config(state="disabled")
        self.results.clear()
        self.errors.clear()
        self.tree.delete(*self.tree.get_children())
        self.status_var.set(
            f"분석 중... (0 / {len(keywords)})  · 첫 검색은 봇 차단 통과로 5~10초 소요"
        )

        threading.Thread(target=self._run_scrape, args=(keywords,), daemon=True).start()

    def _run_scrape(self, keywords: List[str]) -> None:
        total = len(keywords)
        for idx, kw in enumerate(keywords, 1):
            try:
                result = self.scraper.search(kw)
            except Exception as exc:
                message = self._format_error_message(exc)
                self.errors.append(f"{kw}: {message}")
                print(f"[ERROR] {kw}: {message}")
                self.after(0, self._append_error_row, kw, message)
            else:
                self.after(0, self._append_row, result)

            self.after(0, self.status_var.set, f"분석 중... ({idx} / {total})")

        self.after(0, self._on_done, total)

    def _append_row(self, result: KeywordResult) -> None:
        self.results.append(result)
        d = result.to_dict()
        item = self.tree.insert(
            "",
            "end",
            values=(
                d["keyword"],
                d["ttudung"],
                d["total"],
                d["rocket"],
                d["seller_rocket"],
                f'{d["rocket_ratio"]}%',
                f'{d["seller_rocket_ratio"]}%',
                d["seller_rocket_review_over_2000"],
                d["seller_rocket_max_review_count"],
                d["seller_rocket_review_pass"],
                d["rocket_ratio_pass"],
            ),
        )
        self.tree.see(item)

    def _append_error_row(self, keyword: str, message: str) -> None:
        item = self.tree.insert(
            "",
            "end",
            values=(keyword, "실패", f"{message[:120]}", "", "", "", "", "", "", "", ""),
            tags=("error",),
        )
        self.tree.see(item)

    def _on_done(self, count: int) -> None:
        success_count = len(self.results)
        fail_count = len(self.errors)
        self.status_var.set(f"완료 (성공 {success_count}건 / 실패 {fail_count}건)")
        self.run_button.config(state="normal")
        if self.results:
            self.save_button.config(state="normal")
        self._is_running = False

        if self.errors:
            preview = "\n".join(self.errors[:5])
            if len(self.errors) > 5:
                preview += f"\n... 외 {len(self.errors) - 5}건"
            messagebox.showwarning("분석 실패 항목", preview)

    @staticmethod
    def _format_error_message(exc: Exception) -> str:
        raw_message = str(exc).strip()

        if isinstance(exc, TimeoutException):
            return (
                "쿠팡 검색창 또는 상품 목록을 찾지 못했습니다. "
                "자동화 크롬 창을 닫지 말고 잠시 후 다시 시도하세요."
            )

        if isinstance(exc, InvalidSessionIdException):
            return (
                "크롬 자동화 창이 닫혀 세션이 끊겼습니다. "
                "프로그램을 다시 실행하고 크롬 창을 끄지 마세요."
            )

        if isinstance(exc, WebDriverException):
            if "user data directory is already in use" in raw_message.lower():
                return "크롬 프로필이 사용 중입니다. 열린 자동화 크롬 창을 모두 닫고 다시 실행하세요."
            if "invalid session id" in raw_message.lower():
                return (
                    "크롬 자동화 창이 닫혀 세션이 끊겼습니다. "
                    "프로그램을 다시 실행하고 크롬 창을 끄지 마세요."
                )
            if raw_message:
                return raw_message.splitlines()[0]
            return "크롬 자동화 중 오류가 발생했습니다."

        if raw_message:
            return raw_message.splitlines()[0]

        return exc.__class__.__name__

    def on_save(self) -> None:
        if not self.results:
            messagebox.showwarning("저장 불가", "저장할 분석 결과가 없습니다.")
            return

        default_name = f"coupang_keyword_result_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        path = filedialog.asksaveasfilename(
            title="엑셀 저장",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not path:
            return

        try:
            self._save_results_to_xlsx(path)
        except Exception as exc:
            messagebox.showerror("저장 실패", str(exc))
            return

        messagebox.showinfo("저장 완료", f"엑셀 파일을 저장했습니다.\n{path}")

    def _save_results_to_xlsx(self, path: str) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "분석 결과"

        headers = [
            "키워드",
            "뚜둥",
            "전체 상품 수",
            "로켓 개수",
            "판매자로켓 개수",
            "로켓 비율",
            "판매자로켓 비율",
            "댓글 2000+ 상품 수",
            "최대 댓글 수",
            "댓글 기준",
            "로켓 기준",
        ]
        ws.append(headers)

        for result in self._sorted_results_for_export():
            d = result.to_dict()
            ws.append(
                [
                    d["keyword"],
                    d["ttudung"],
                    d["total"],
                    d["rocket"],
                    d["seller_rocket"],
                    d["rocket_ratio"] / 100,
                    d["seller_rocket_ratio"] / 100,
                    d["seller_rocket_review_over_2000"],
                    d["seller_rocket_max_review_count"],
                    d["seller_rocket_review_pass"],
                    d["rocket_ratio_pass"],
                ]
            )

        self._format_worksheet(ws)
        wb.save(path)

    def _sorted_results_for_export(self) -> List[KeywordResult]:
        return sorted(
            self.results,
            key=lambda result: (
                self.TTUDUNG_SORT_ORDER.get(result.ttudung, 99),
                result.rocket_ratio,
                result.seller_rocket_review_over_2000,
                result.seller_rocket_max_review_count,
                result.keyword,
            ),
        )

    @staticmethod
    def _format_worksheet(ws) -> None:
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")
            row[0].alignment = Alignment(horizontal="left")

        for col_idx in (6, 7):
            for cell in ws.iter_cols(
                min_col=col_idx,
                max_col=col_idx,
                min_row=2,
            ):
                for data_cell in cell:
                    data_cell.number_format = "0.0%"

        widths = [24, 10, 12, 10, 14, 10, 14, 16, 12, 10, 10]
        for idx, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def _on_close(self) -> None:
        try:
            self.scraper.close()
        finally:
            self.destroy()


def run() -> None:
    app = App()
    app.mainloop()
